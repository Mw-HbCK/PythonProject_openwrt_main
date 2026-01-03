#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
报表生成服务
处理报表数据收集、图表生成、格式转换等操作
"""
import os
import sys
import json
import base64
from datetime import datetime, timedelta
from typing import Dict, List, Optional, Tuple, Any

# 添加项目根目录到路径
project_root = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
if project_root not in sys.path:
    sys.path.insert(0, project_root)

from app.models.database_models import db, Device, TotalTraffic, DeviceTraffic
from app.models.alert_models import AlertHistory
from sqlalchemy import desc


class ReportService:
    """报表生成服务类"""
    
    def __init__(self, app=None):
        """
        初始化报表生成服务
        
        Args:
            app: Flask应用实例（可选）
        """
        self.app = app
    
    def generate_report(self, report_type: str, period_start: datetime, period_end: datetime, 
                       report_config: Dict[str, Any]) -> Tuple[Dict[str, str], int]:
        """
        生成报表主方法
        
        Args:
            report_type: 报表类型（daily/weekly/monthly）
            period_start: 报表周期开始时间
            period_end: 报表周期结束时间
            report_config: 报表配置（包含格式、存储目录等）
            
        Returns:
            tuple: (文件路径字典, 总文件大小)
        """
        try:
            # 确保报表目录存在
            report_dir = report_config.get('report_dir', './reports')
            # 将相对路径转换为绝对路径（基于项目根目录）
            if not os.path.isabs(report_dir):
                # 获取项目根目录
                current_file = os.path.abspath(__file__)
                project_root = os.path.dirname(os.path.dirname(os.path.dirname(current_file)))
                # 处理相对路径，移除开头的./
                if report_dir.startswith('./'):
                    report_dir = report_dir[2:]
                report_dir = os.path.join(project_root, report_dir)
            os.makedirs(report_dir, exist_ok=True)
            
            # 生成报表文件名前缀
            period_str = period_start.strftime('%Y%m%d')
            if report_type == 'weekly':
                # 周报：使用周期开始日期
                period_str = period_start.strftime('%Y%m%d')
            elif report_type == 'monthly':
                # 月报：使用年月
                period_str = period_start.strftime('%Y%m')
            
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            file_prefix = f"report_{report_type}_{period_str}_{timestamp}"
            
            # 1. 收集报表数据
            report_data = self._collect_report_data(period_start, period_end)
            
            # 2. 生成图表图片
            chart_paths = {}
            output_dir = os.path.join(report_dir, file_prefix + '_charts')
            os.makedirs(output_dir, exist_ok=True)
            
            try:
                chart_paths = self._generate_chart_images(report_data, period_start, period_end, output_dir)
            except Exception as e:
                print(f"[报表服务] 生成图表失败: {e}", file=sys.stderr)
                # 图表生成失败不影响报表生成，继续执行
            
            # 3. 生成各种格式的报表
            file_paths = {}
            total_size = 0
            
            # 生成HTML报表
            if report_config.get('generate_html', 'true').lower() == 'true':
                try:
                    html_path = os.path.join(report_dir, f"{file_prefix}.html")
                    self._generate_html_report(report_data, chart_paths, html_path, period_start, period_end, report_type)
                    file_paths['html'] = html_path
                    total_size += os.path.getsize(html_path)
                except Exception as e:
                    print(f"[报表服务] 生成HTML报表失败: {e}", file=sys.stderr)
            
            # 生成PDF报表
            if report_config.get('generate_pdf', 'true').lower() == 'true':
                try:
                    pdf_path = os.path.join(report_dir, f"{file_prefix}.pdf")
                    pdf_library = report_config.get('pdf_library', 'reportlab')
                    self._generate_pdf_report(report_data, chart_paths, pdf_path, period_start, period_end, report_type, pdf_library)
                    file_paths['pdf'] = pdf_path
                    total_size += os.path.getsize(pdf_path)
                except Exception as e:
                    print(f"[报表服务] 生成PDF报表失败: {e}", file=sys.stderr)
            
            # 生成Excel报表
            if report_config.get('generate_excel', 'true').lower() == 'true':
                try:
                    excel_path = os.path.join(report_dir, f"{file_prefix}.xlsx")
                    self._generate_excel_report(report_data, chart_paths, excel_path, period_start, period_end, report_type)
                    file_paths['excel'] = excel_path
                    total_size += os.path.getsize(excel_path)
                except Exception as e:
                    print(f"[报表服务] 生成Excel报表失败: {e}", file=sys.stderr)
            
            # 清理临时图表目录
            try:
                if os.path.exists(output_dir):
                    import shutil
                    shutil.rmtree(output_dir)
            except Exception as e:
                print(f"[报表服务] 清理临时目录失败: {e}", file=sys.stderr)
            
            return file_paths, total_size
            
        except Exception as e:
            raise Exception(f"生成报表失败: {str(e)}")
    
    def _collect_report_data(self, period_start: datetime, period_end: datetime) -> Dict[str, Any]:
        """
        收集报表数据
        复用stats API的逻辑
        
        Args:
            period_start: 开始时间
            period_end: 结束时间
            
        Returns:
            dict: 报表数据字典
        """
        try:
            # 1. 流量统计摘要
            traffic_summary = self._calculate_traffic_summary(period_start, period_end)
            
            # 2. 峰值速率统计
            peak_speed = self._get_peak_speed(period_start, period_end)
            
            # 3. 设备排名Top 10
            device_ranking = self._get_device_ranking(limit=10, start_time=period_start, end_time=period_end)
            
            # 4. 告警事件统计
            alert_stats = self._get_alert_statistics(period_start, period_end)
            
            # 5. 流量趋势数据（用于图表）
            traffic_trend = self._get_traffic_trend(period_start, period_end)
            
            return {
                'traffic_summary': traffic_summary,
                'peak_speed': peak_speed,
                'device_ranking': device_ranking,
                'alert_stats': alert_stats,
                'traffic_trend': traffic_trend,
                'period_start': period_start.isoformat(),
                'period_end': period_end.isoformat()
            }
        except Exception as e:
            print(f"[报表服务] 收集报表数据失败: {e}", file=sys.stderr)
            import traceback
            traceback.print_exc()
            return {
                'traffic_summary': {},
                'peak_speed': {},
                'device_ranking': [],
                'alert_stats': {},
                'traffic_trend': [],
                'period_start': period_start.isoformat(),
                'period_end': period_end.isoformat()
            }
    
    def _calculate_traffic_summary(self, start_time: datetime, end_time: datetime) -> Dict[str, Any]:
        """计算流量汇总（复用database_api的逻辑）"""
        try:
            query = TotalTraffic.query.filter(
                TotalTraffic.timestamp >= start_time,
                TotalTraffic.timestamp <= end_time
            )
            
            # 获取首尾记录以计算流量增量
            first_record = query.order_by(TotalTraffic.timestamp.asc()).first()
            last_record = query.order_by(TotalTraffic.timestamp.desc()).first()
            
            # 计算总流量增量
            total_download = 0
            total_upload = 0
            if first_record and last_record:
                total_download = max(0, last_record.total_download_bytes - first_record.total_download_bytes)
                total_upload = max(0, last_record.total_upload_bytes - first_record.total_upload_bytes)
            
            # 计算平均速率
            records = query.all()
            avg_down_speed = 0
            avg_up_speed = 0
            if records:
                avg_down_speed = sum(r.down_speed_bytes for r in records) / len(records)
                avg_up_speed = sum(r.up_speed_bytes for r in records) / len(records)
            
            return {
                'total_download_bytes': total_download,
                'total_upload_bytes': total_upload,
                'total_traffic_bytes': total_download + total_upload,
                'total_download_formatted': self._format_bytes_to_size(total_download),
                'total_upload_formatted': self._format_bytes_to_size(total_upload),
                'total_traffic_formatted': self._format_bytes_to_size(total_download + total_upload),
                'avg_down_speed_bytes': avg_down_speed,
                'avg_up_speed_bytes': avg_up_speed,
                'avg_down_speed_formatted': self._format_bytes_to_speed(avg_down_speed),
                'avg_up_speed_formatted': self._format_bytes_to_speed(avg_up_speed),
                'record_count': len(records)
            }
        except Exception as e:
            return {
                'total_download_bytes': 0,
                'total_upload_bytes': 0,
                'total_traffic_bytes': 0,
                'total_download_formatted': '0 B',
                'total_upload_formatted': '0 B',
                'total_traffic_formatted': '0 B',
                'avg_down_speed_bytes': 0,
                'avg_up_speed_bytes': 0,
                'avg_down_speed_formatted': '0 B/s',
                'avg_up_speed_formatted': '0 B/s',
                'record_count': 0
            }
    
    def _get_peak_speed(self, start_time: datetime, end_time: datetime) -> Dict[str, Any]:
        """获取峰值速率"""
        try:
            query = TotalTraffic.query.filter(
                TotalTraffic.timestamp >= start_time,
                TotalTraffic.timestamp <= end_time
            )
            
            # 获取下行峰值
            peak_down_record = query.order_by(desc(TotalTraffic.down_speed_bytes)).first()
            # 获取上行峰值
            peak_up_record = query.order_by(desc(TotalTraffic.up_speed_bytes)).first()
            
            result = {
                'peak_down_speed_bytes': 0,
                'peak_down_speed_formatted': '0 B/s',
                'peak_down_speed_time': None,
                'peak_up_speed_bytes': 0,
                'peak_up_speed_formatted': '0 B/s',
                'peak_up_speed_time': None
            }
            
            if peak_down_record:
                result['peak_down_speed_bytes'] = peak_down_record.down_speed_bytes
                result['peak_down_speed_formatted'] = self._format_bytes_to_speed(peak_down_record.down_speed_bytes)
                result['peak_down_speed_time'] = peak_down_record.timestamp.isoformat() if peak_down_record.timestamp else None
            
            if peak_up_record:
                result['peak_up_speed_bytes'] = peak_up_record.up_speed_bytes
                result['peak_up_speed_formatted'] = self._format_bytes_to_speed(peak_up_record.up_speed_bytes)
                result['peak_up_speed_time'] = peak_up_record.timestamp.isoformat() if peak_up_record.timestamp else None
            
            return result
        except Exception as e:
            return {
                'peak_down_speed_bytes': 0,
                'peak_down_speed_formatted': '0 B/s',
                'peak_down_speed_time': None,
                'peak_up_speed_bytes': 0,
                'peak_up_speed_formatted': '0 B/s',
                'peak_up_speed_time': None
            }
    
    def _get_device_ranking(self, limit: int = 10, start_time: Optional[datetime] = None, 
                           end_time: Optional[datetime] = None) -> List[Dict[str, Any]]:
        """获取设备排名"""
        try:
            devices = Device.query.all()
            ranking = []
            
            for device in devices:
                query = DeviceTraffic.query.filter_by(device_id=device.id)
                
                if start_time:
                    query = query.filter(DeviceTraffic.timestamp >= start_time)
                if end_time:
                    query = query.filter(DeviceTraffic.timestamp <= end_time)
                
                # 获取首尾记录计算流量增量
                first_record = query.order_by(DeviceTraffic.timestamp.asc()).first()
                last_record = query.order_by(DeviceTraffic.timestamp.desc()).first()
                
                if first_record and last_record:
                    total_download = max(0, last_record.total_download_bytes - first_record.total_download_bytes)
                    total_upload = max(0, last_record.total_upload_bytes - first_record.total_upload_bytes)
                    total_traffic = total_download + total_upload
                    
                    ranking.append({
                        'device_id': device.id,
                        'device_name': device.hostname or '未知',
                        'mac': device.mac,
                        'ip': device.ip or '-',
                        'total_download_bytes': total_download,
                        'total_upload_bytes': total_upload,
                        'total_traffic_bytes': total_traffic,
                        'total_download_formatted': self._format_bytes_to_size(total_download),
                        'total_upload_formatted': self._format_bytes_to_size(total_upload),
                        'total_traffic_formatted': self._format_bytes_to_size(total_traffic)
                    })
            
            # 按总流量排序
            ranking.sort(key=lambda x: x['total_traffic_bytes'], reverse=True)
            
            return ranking[:limit]
        except Exception as e:
            return []
    
    def _get_alert_statistics(self, period_start: datetime, period_end: datetime) -> Dict[str, Any]:
        """获取告警事件统计"""
        try:
            # 查询指定时间段内的告警记录
            alerts = AlertHistory.query.filter(
                AlertHistory.triggered_at >= period_start,
                AlertHistory.triggered_at <= period_end
            ).all()
            
            # 按类型统计
            alert_by_type = {}
            alert_by_severity = {'critical': 0, 'warning': 0, 'info': 0}
            
            for alert in alerts:
                # 按类型统计
                alert_type = alert.alert_type
                alert_by_type[alert_type] = alert_by_type.get(alert_type, 0) + 1
                
                # 按严重程度统计
                severity = alert.severity
                if severity in alert_by_severity:
                    alert_by_severity[severity] += 1
            
            return {
                'total_count': len(alerts),
                'by_type': alert_by_type,
                'by_severity': alert_by_severity,
                'alerts': [alert.to_dict() for alert in alerts[:50]]  # 最多返回50条
            }
        except Exception as e:
            return {
                'total_count': 0,
                'by_type': {},
                'by_severity': {'critical': 0, 'warning': 0, 'info': 0},
                'alerts': []
            }
    
    def _get_traffic_trend(self, period_start: datetime, period_end: datetime) -> List[Dict[str, Any]]:
        """获取流量趋势数据（用于图表）"""
        try:
            # 获取时间段内的所有流量记录
            records = TotalTraffic.query.filter(
                TotalTraffic.timestamp >= period_start,
                TotalTraffic.timestamp <= period_end
            ).order_by(TotalTraffic.timestamp.asc()).all()
            
            # 按小时聚合数据（减少数据点）
            trend_data = []
            current_hour = None
            hour_data = []
            
            for record in records:
                record_hour = record.timestamp.replace(minute=0, second=0, microsecond=0)
                
                if current_hour is None or record_hour != current_hour:
                    # 保存上一个小时的数据
                    if hour_data:
                        trend_data.append({
                            'timestamp': current_hour.isoformat(),
                            'down_speed': sum(r.down_speed_bytes for r in hour_data) // len(hour_data),
                            'up_speed': sum(r.up_speed_bytes for r in hour_data) // len(hour_data),
                            'total_download': max(r.total_download_bytes for r in hour_data),
                            'total_upload': max(r.total_upload_bytes for r in hour_data)
                        })
                    
                    # 开始新的小时
                    current_hour = record_hour
                    hour_data = [record]
                else:
                    hour_data.append(record)
            
            # 处理最后一个小时
            if hour_data:
                trend_data.append({
                    'timestamp': current_hour.isoformat(),
                    'down_speed': sum(r.down_speed_bytes for r in hour_data) // len(hour_data),
                    'up_speed': sum(r.up_speed_bytes for r in hour_data) // len(hour_data),
                    'total_download': max(r.total_download_bytes for r in hour_data),
                    'total_upload': max(r.total_upload_bytes for r in hour_data)
                })
            
            return trend_data
        except Exception as e:
            return []
    
    def _generate_chart_images(self, report_data: Dict[str, Any], period_start: datetime, 
                               period_end: datetime, output_dir: str) -> Dict[str, str]:
        """
        生成图表图片
        使用matplotlib生成流量趋势图和设备排名图
        
        Args:
            report_data: 报表数据
            period_start: 开始时间
            period_end: 结束时间
            output_dir: 输出目录
            
        Returns:
            dict: 图表文件路径字典
        """
        chart_paths = {}
        
        try:
            import matplotlib
            matplotlib.use('Agg')  # 使用非交互式后端
            import matplotlib.pyplot as plt
            import matplotlib.dates as mdates
            from matplotlib import font_manager
            
            # 设置中文字体（尝试使用系统字体）
            try:
                # Windows系统
                if sys.platform == 'win32':
                    font_paths = [
                        'C:/Windows/Fonts/simhei.ttf',  # 黑体
                        'C:/Windows/Fonts/simsun.ttc',   # 宋体
                        'C:/Windows/Fonts/msyh.ttc',     # 微软雅黑
                    ]
                else:
                    # Linux/Mac系统
                    font_paths = [
                        '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf',
                        '/System/Library/Fonts/PingFang.ttc',  # macOS
                    ]
                
                font_set = False
                for font_path in font_paths:
                    if os.path.exists(font_path):
                        prop = font_manager.FontProperties(fname=font_path)
                        plt.rcParams['font.sans-serif'] = [prop.get_name()]
                        plt.rcParams['axes.unicode_minus'] = False
                        font_set = True
                        break
                
                if not font_set:
                    # 如果没有找到中文字体，使用默认字体
                    plt.rcParams['font.sans-serif'] = ['Arial', 'DejaVu Sans']
                    plt.rcParams['axes.unicode_minus'] = False
            except Exception as e:
                print(f"[报表服务] 设置中文字体失败，使用默认字体: {e}", file=sys.stderr)
                plt.rcParams['font.sans-serif'] = ['Arial', 'DejaVu Sans']
                plt.rcParams['axes.unicode_minus'] = False
            
            # 1. 生成流量趋势图
            try:
                traffic_trend = report_data.get('traffic_trend', [])
                if traffic_trend:
                    fig, ax = plt.subplots(figsize=(12, 6))
                    
                    timestamps = [datetime.fromisoformat(d['timestamp']) for d in traffic_trend]
                    down_speeds = [d['down_speed'] / (1024 * 1024) for d in traffic_trend]  # 转换为MB/s
                    up_speeds = [d['up_speed'] / (1024 * 1024) for d in traffic_trend]
                    
                    ax.plot(timestamps, down_speeds, label='下行速率', color='#3498db', linewidth=2)
                    ax.plot(timestamps, up_speeds, label='上行速率', color='#e74c3c', linewidth=2)
                    
                    ax.set_xlabel('时间', fontsize=12)
                    ax.set_ylabel('速率 (MB/s)', fontsize=12)
                    ax.set_title('流量趋势图', fontsize=14, fontweight='bold')
                    ax.legend(loc='best')
                    ax.grid(True, alpha=0.3)
                    
                    # 格式化x轴日期
                    ax.xaxis.set_major_formatter(mdates.DateFormatter('%m-%d %H:%M'))
                    plt.xticks(rotation=45)
                    
                    plt.tight_layout()
                    
                    trend_path = os.path.join(output_dir, 'traffic_trend.png')
                    plt.savefig(trend_path, dpi=300, bbox_inches='tight')
                    plt.close()
                    chart_paths['traffic_trend'] = trend_path
            except Exception as e:
                print(f"[报表服务] 生成流量趋势图失败: {e}", file=sys.stderr)
            
            # 2. 生成设备排名图
            try:
                device_ranking = report_data.get('device_ranking', [])
                if device_ranking:
                    fig, ax = plt.subplots(figsize=(12, 6))
                    
                    device_names = [d['device_name'][:15] for d in device_ranking]  # 限制名称长度
                    traffic_values = [d['total_traffic_bytes'] / (1024 * 1024 * 1024) for d in device_ranking]  # 转换为GB
                    
                    colors = plt.cm.viridis(range(len(device_names)))
                    bars = ax.barh(range(len(device_names)), traffic_values, color=colors)
                    
                    ax.set_yticks(range(len(device_names)))
                    ax.set_yticklabels(device_names)
                    ax.set_xlabel('总流量 (GB)', fontsize=12)
                    ax.set_title('设备流量排名 Top 10', fontsize=14, fontweight='bold')
                    ax.grid(True, alpha=0.3, axis='x')
                    
                    # 在柱状图上显示数值
                    for i, (bar, value) in enumerate(zip(bars, traffic_values)):
                        ax.text(value, i, f'{value:.2f} GB', va='center', ha='left', fontsize=9)
                    
                    plt.tight_layout()
                    
                    ranking_path = os.path.join(output_dir, 'device_ranking.png')
                    plt.savefig(ranking_path, dpi=300, bbox_inches='tight')
                    plt.close()
                    chart_paths['device_ranking'] = ranking_path
            except Exception as e:
                print(f"[报表服务] 生成设备排名图失败: {e}", file=sys.stderr)
            
        except ImportError:
            print("[报表服务] matplotlib未安装，跳过图表生成", file=sys.stderr)
        except Exception as e:
            print(f"[报表服务] 生成图表失败: {e}", file=sys.stderr)
            import traceback
            traceback.print_exc()
        
        return chart_paths
    
    def _generate_html_report(self, report_data: Dict[str, Any], chart_paths: Dict[str, str],
                            output_path: str, period_start: datetime, period_end: datetime, 
                            report_type: str):
        """
        生成HTML报表
        使用Jinja2模板
        
        Args:
            report_data: 报表数据
            chart_paths: 图表文件路径字典
            output_path: 输出文件路径
            period_start: 开始时间
            period_end: 结束时间
            report_type: 报表类型
        """
        try:
            from flask import render_template
            
            # 将图表转换为base64嵌入（如果图表存在）
            chart_base64 = {}
            for chart_name, chart_path in chart_paths.items():
                if os.path.exists(chart_path):
                    try:
                        with open(chart_path, 'rb') as f:
                            chart_data = base64.b64encode(f.read()).decode('utf-8')
                            chart_base64[chart_name] = f"data:image/png;base64,{chart_data}"
                    except Exception as e:
                        print(f"[报表服务] 读取图表文件失败 {chart_path}: {e}", file=sys.stderr)
            
            # 渲染模板
            html_content = render_template(
                'reports/report_template.html',
                report_data=report_data,
                chart_base64=chart_base64,
                period_start=period_start,
                period_end=period_end,
                report_type=report_type
            )
            
            # 保存HTML文件
            with open(output_path, 'w', encoding='utf-8') as f:
                f.write(html_content)
                
        except Exception as e:
            raise Exception(f"生成HTML报表失败: {str(e)}")
    
    def _generate_pdf_report(self, report_data: Dict[str, Any], chart_paths: Dict[str, str],
                            output_path: str, period_start: datetime, period_end: datetime,
                            report_type: str, pdf_library: str = 'reportlab'):
        """
        生成PDF报表
        
        Args:
            report_data: 报表数据
            chart_paths: 图表文件路径字典
            output_path: 输出文件路径
            period_start: 开始时间
            period_end: 结束时间
            report_type: 报表类型
            pdf_library: PDF生成库（reportlab/weasyprint）
        """
        if pdf_library == 'reportlab':
            self._generate_pdf_reportlab(report_data, chart_paths, output_path, period_start, period_end, report_type)
        elif pdf_library == 'weasyprint':
            self._generate_pdf_weasyprint(report_data, chart_paths, output_path, period_start, period_end, report_type)
        else:
            raise ValueError(f"不支持的PDF库: {pdf_library}")
    
    def _generate_pdf_reportlab(self, report_data: Dict[str, Any], chart_paths: Dict[str, str],
                               output_path: str, period_start: datetime, period_end: datetime,
                               report_type: str):
        """使用reportlab生成PDF"""
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.lib.units import cm
            from reportlab.lib import colors
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image, PageBreak
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.enums import TA_CENTER, TA_LEFT
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont
            
            # 注册中文字体
            chinese_font_name = 'ChineseFont'
            chinese_font_bold_name = 'ChineseFontBold'
            font_registered = False
            bold_font_registered = False
            
            # 尝试注册中文字体（Windows常见路径）
            if sys.platform == 'win32':
                # 普通字体路径列表
                normal_font_paths = [
                    'C:/Windows/Fonts/simsun.ttc',  # 宋体
                    'C:/Windows/Fonts/msyh.ttc',    # 微软雅黑
                    'C:/Windows/Fonts/simkai.ttf',  # 楷体
                ]
                # 粗体字体路径列表
                bold_font_paths = [
                    'C:/Windows/Fonts/simhei.ttf',  # 黑体（粗体）
                    'C:/Windows/Fonts/msyhbd.ttc',  # 微软雅黑粗体
                ]
            else:
                # Linux/Mac字体路径
                normal_font_paths = [
                    '/usr/share/fonts/truetype/wqy/wqy-microhei.ttc',
                    '/usr/share/fonts/truetype/arphic/uming.ttc',
                    '/System/Library/Fonts/PingFang.ttc',  # macOS
                ]
                bold_font_paths = [
                    '/usr/share/fonts/truetype/wqy/wqy-zenhei.ttc',
                ]
            
            # 注册普通字体
            for font_path in normal_font_paths:
                if os.path.exists(font_path):
                    try:
                        pdfmetrics.registerFont(TTFont(chinese_font_name, font_path))
                        font_registered = True
                        break
                    except Exception as e:
                        print(f"[报表服务] 注册字体失败 {font_path}: {e}", file=sys.stderr)
                        continue
            
            # 注册粗体字体
            for font_path in bold_font_paths:
                if os.path.exists(font_path):
                    try:
                        pdfmetrics.registerFont(TTFont(chinese_font_bold_name, font_path))
                        bold_font_registered = True
                        break
                    except Exception as e:
                        print(f"[报表服务] 注册粗体字体失败 {font_path}: {e}", file=sys.stderr)
                        continue
            
            # 如果没有找到粗体字体，使用普通字体作为粗体
            if font_registered and not bold_font_registered:
                try:
                    # 使用普通字体路径注册粗体
                    for font_path in normal_font_paths:
                        if os.path.exists(font_path):
                            pdfmetrics.registerFont(TTFont(chinese_font_bold_name, font_path))
                            bold_font_registered = True
                            break
                except Exception as e:
                    print(f"[报表服务] 使用普通字体作为粗体失败: {e}", file=sys.stderr)
            
            if not font_registered:
                print("[报表服务] 警告：未找到中文字体，PDF中的中文可能无法正常显示", file=sys.stderr)
                # 使用默认字体作为后备
                chinese_font_name = 'Helvetica'
                chinese_font_bold_name = 'Helvetica-Bold'
            
            # 创建PDF文档
            doc = SimpleDocTemplate(output_path, pagesize=A4)
            story = []
            
            # 定义样式（使用中文字体）
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontName=chinese_font_bold_name,
                fontSize=24,
                textColor=colors.HexColor('#2c3e50'),
                spaceAfter=30,
                alignment=TA_CENTER
            )
            
            heading_style = ParagraphStyle(
                'CustomHeading',
                parent=styles['Heading2'],
                fontName=chinese_font_bold_name,
                fontSize=16,
                textColor=colors.HexColor('#34495e'),
                spaceAfter=12,
                spaceBefore=12
            )
            
            normal_style = ParagraphStyle(
                'CustomNormal',
                parent=styles['Normal'],
                fontName=chinese_font_name,
                fontSize=10
            )
            
            # 标题
            report_type_names = {'daily': '日报', 'weekly': '周报', 'monthly': '月报'}
            title = f"{report_type_names.get(report_type, '报表')} - {period_start.strftime('%Y-%m-%d')}"
            story.append(Paragraph(title, title_style))
            story.append(Spacer(1, 0.5*cm))
            
            # 流量统计摘要
            story.append(Paragraph("流量统计摘要", heading_style))
            traffic_summary = report_data.get('traffic_summary', {})
            summary_data = [
                ['指标', '数值'],
                ['总下载流量', traffic_summary.get('total_download_formatted', '0 B')],
                ['总上传流量', traffic_summary.get('total_upload_formatted', '0 B')],
                ['总流量', traffic_summary.get('total_traffic_formatted', '0 B')],
                ['平均下行速率', traffic_summary.get('avg_down_speed_formatted', '0 B/s')],
                ['平均上行速率', traffic_summary.get('avg_up_speed_formatted', '0 B/s')],
            ]
            # 将表格数据转换为Paragraph以支持中文字体
            summary_data_para = []
            for row in summary_data:
                para_row = []
                for cell in row:
                    if isinstance(cell, str):
                        para_row.append(Paragraph(cell, normal_style))
                    else:
                        para_row.append(cell)
                summary_data_para.append(para_row)
            
            summary_table = Table(summary_data_para, colWidths=[6*cm, 6*cm])
            summary_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3498db')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), chinese_font_bold_name),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))
            story.append(summary_table)
            story.append(Spacer(1, 0.5*cm))
            
            # 设备排名
            story.append(Paragraph("设备流量排名 Top 10", heading_style))
            device_ranking = report_data.get('device_ranking', [])
            if device_ranking:
                ranking_data = [['排名', '设备名称', 'MAC地址', '总流量']]
                for i, device in enumerate(device_ranking[:10], 1):
                    ranking_data.append([
                        str(i),
                        device.get('device_name', '未知')[:20],
                        device.get('mac', '-')[:17],
                        device.get('total_traffic_formatted', '0 B')
                    ])
                
                # 将表格数据转换为Paragraph以支持中文字体
                ranking_data_para = []
                for row in ranking_data:
                    para_row = []
                    for cell in row:
                        if isinstance(cell, str):
                            para_row.append(Paragraph(cell, normal_style))
                        else:
                            para_row.append(cell)
                    ranking_data_para.append(para_row)
                
                ranking_table = Table(ranking_data_para, colWidths=[1*cm, 4*cm, 4*cm, 3*cm])
                ranking_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#e74c3c')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, 0), chinese_font_bold_name),
                    ('FONTSIZE', (0, 0), (-1, 0), 11),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.lightgrey),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ]))
                story.append(ranking_table)
            else:
                story.append(Paragraph("暂无数据", normal_style))
            
            story.append(Spacer(1, 0.5*cm))
            
            # 插入图表
            if 'traffic_trend' in chart_paths and os.path.exists(chart_paths['traffic_trend']):
                story.append(Paragraph("流量趋势图", heading_style))
                img = Image(chart_paths['traffic_trend'], width=16*cm, height=8*cm)
                story.append(img)
                story.append(Spacer(1, 0.3*cm))
            
            if 'device_ranking' in chart_paths and os.path.exists(chart_paths['device_ranking']):
                story.append(Paragraph("设备排名图", heading_style))
                img = Image(chart_paths['device_ranking'], width=16*cm, height=8*cm)
                story.append(img)
                story.append(Spacer(1, 0.3*cm))
            
            # 告警统计
            story.append(Paragraph("告警事件统计", heading_style))
            alert_stats = report_data.get('alert_stats', {})
            alert_data = [
                ['指标', '数值'],
                ['总告警数', str(alert_stats.get('total_count', 0))],
                ['严重告警', str(alert_stats.get('by_severity', {}).get('critical', 0))],
                ['警告', str(alert_stats.get('by_severity', {}).get('warning', 0))],
                ['信息', str(alert_stats.get('by_severity', {}).get('info', 0))],
            ]
            
            # 将表格数据转换为Paragraph以支持中文字体
            alert_data_para = []
            for row in alert_data:
                para_row = []
                for cell in row:
                    if isinstance(cell, str):
                        para_row.append(Paragraph(cell, normal_style))
                    else:
                        para_row.append(cell)
                alert_data_para.append(para_row)
            
            alert_table = Table(alert_data_para, colWidths=[6*cm, 6*cm])
            alert_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f39c12')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), chinese_font_bold_name),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))
            story.append(alert_table)
            
            # 生成PDF
            doc.build(story)
            
        except ImportError:
            raise Exception("reportlab未安装，无法生成PDF报表")
        except Exception as e:
            raise Exception(f"使用reportlab生成PDF失败: {str(e)}")
    
    def _generate_pdf_weasyprint(self, report_data: Dict[str, Any], chart_paths: Dict[str, str],
                                 output_path: str, period_start: datetime, period_end: datetime,
                                 report_type: str):
        """使用weasyprint生成PDF"""
        try:
            from flask import render_template
            from weasyprint import HTML
            
            # 将图表转换为base64嵌入
            chart_base64 = {}
            for chart_name, chart_path in chart_paths.items():
                if os.path.exists(chart_path):
                    try:
                        with open(chart_path, 'rb') as f:
                            chart_data = base64.b64encode(f.read()).decode('utf-8')
                            chart_base64[chart_name] = f"data:image/png;base64,{chart_data}"
                    except Exception as e:
                        print(f"[报表服务] 读取图表文件失败 {chart_path}: {e}", file=sys.stderr)
            
            # 渲染HTML模板
            html_content = render_template(
                'reports/report_template_pdf.html',
                report_data=report_data,
                chart_base64=chart_base64,
                period_start=period_start,
                period_end=period_end,
                report_type=report_type
            )
            
            # 转换为PDF
            HTML(string=html_content).write_pdf(output_path)
            
        except ImportError:
            raise Exception("weasyprint未安装，无法生成PDF报表")
        except Exception as e:
            raise Exception(f"使用weasyprint生成PDF失败: {str(e)}")
    
    def _generate_excel_report(self, report_data: Dict[str, Any], chart_paths: Dict[str, str],
                               output_path: str, period_start: datetime, period_end: datetime,
                               report_type: str):
        """
        生成Excel报表
        使用openpyxl库
        
        Args:
            report_data: 报表数据
            chart_paths: 图表文件路径字典
            output_path: 输出文件路径
            period_start: 开始时间
            period_end: 结束时间
            report_type: 报表类型
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.chart import LineChart, BarChart, Reference
            from openpyxl.drawing.image import Image as ExcelImage
            
            wb = Workbook()
            ws = wb.active
            ws.title = "报表摘要"
            
            # 标题样式
            title_font = Font(name='Arial', size=16, bold=True)
            header_fill = PatternFill(start_color="3498db", end_color="3498db", fill_type="solid")
            header_font = Font(name='Arial', size=12, bold=True, color="FFFFFF")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # 标题
            report_type_names = {'daily': '日报', 'weekly': '周报', 'monthly': '月报'}
            title = f"{report_type_names.get(report_type, '报表')} - {period_start.strftime('%Y-%m-%d')}"
            ws['A1'] = title
            ws['A1'].font = title_font
            ws.merge_cells('A1:D1')
            ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
            
            row = 3
            
            # 流量统计摘要
            ws[f'A{row}'] = "流量统计摘要"
            ws[f'A{row}'].font = Font(name='Arial', size=14, bold=True)
            row += 1
            
            summary_headers = ['指标', '数值']
            for col, header in enumerate(summary_headers, 1):
                cell = ws.cell(row=row, column=col)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            row += 1
            traffic_summary = report_data.get('traffic_summary', {})
            summary_data = [
                ['总下载流量', traffic_summary.get('total_download_formatted', '0 B')],
                ['总上传流量', traffic_summary.get('total_upload_formatted', '0 B')],
                ['总流量', traffic_summary.get('total_traffic_formatted', '0 B')],
                ['平均下行速率', traffic_summary.get('avg_down_speed_formatted', '0 B/s')],
                ['平均上行速率', traffic_summary.get('avg_up_speed_formatted', '0 B/s')],
            ]
            
            for data_row in summary_data:
                for col, value in enumerate(data_row, 1):
                    cell = ws.cell(row=row, column=col)
                    cell.value = value
                    cell.border = border
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                row += 1
            
            row += 1
            
            # 设备排名
            ws[f'A{row}'] = "设备流量排名 Top 10"
            ws[f'A{row}'].font = Font(name='Arial', size=14, bold=True)
            row += 1
            
            device_ranking = report_data.get('device_ranking', [])
            if device_ranking:
                ranking_headers = ['排名', '设备名称', 'MAC地址', '总流量']
                for col, header in enumerate(ranking_headers, 1):
                    cell = ws.cell(row=row, column=col)
                    cell.value = header
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                
                row += 1
                for i, device in enumerate(device_ranking[:10], 1):
                    ws.cell(row=row, column=1).value = i
                    ws.cell(row=row, column=2).value = device.get('device_name', '未知')
                    ws.cell(row=row, column=3).value = device.get('mac', '-')
                    ws.cell(row=row, column=4).value = device.get('total_traffic_formatted', '0 B')
                    
                    for col in range(1, 5):
                        cell = ws.cell(row=row, column=col)
                        cell.border = border
                        cell.alignment = Alignment(horizontal='left', vertical='center')
                    row += 1
            
            row += 1
            
            # 告警统计
            ws[f'A{row}'] = "告警事件统计"
            ws[f'A{row}'].font = Font(name='Arial', size=14, bold=True)
            row += 1
            
            alert_stats = report_data.get('alert_stats', {})
            alert_headers = ['指标', '数值']
            for col, header in enumerate(alert_headers, 1):
                cell = ws.cell(row=row, column=col)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            row += 1
            alert_data = [
                ['总告警数', alert_stats.get('total_count', 0)],
                ['严重告警', alert_stats.get('by_severity', {}).get('critical', 0)],
                ['警告', alert_stats.get('by_severity', {}).get('warning', 0)],
                ['信息', alert_stats.get('by_severity', {}).get('info', 0)],
            ]
            
            for data_row in alert_data:
                for col, value in enumerate(data_row, 1):
                    cell = ws.cell(row=row, column=col)
                    cell.value = value
                    cell.border = border
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                row += 1
            
            # 调整列宽
            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 30
            ws.column_dimensions['C'].width = 20
            ws.column_dimensions['D'].width = 15
            
            # 保存Excel文件
            wb.save(output_path)
            
        except ImportError:
            raise Exception("openpyxl未安装，无法生成Excel报表")
        except Exception as e:
            raise Exception(f"生成Excel报表失败: {str(e)}")
    
    @staticmethod
    def _format_bytes_to_size(bytes_value):
        """将字节转换为可读的大小字符串"""
        if bytes_value < 1024:
            return f"{bytes_value} B"
        elif bytes_value < 1024 * 1024:
            return f"{bytes_value / 1024:.2f} KB"
        elif bytes_value < 1024 * 1024 * 1024:
            return f"{bytes_value / (1024 * 1024):.2f} MB"
        elif bytes_value < 1024 * 1024 * 1024 * 1024:
            return f"{bytes_value / (1024 * 1024 * 1024):.2f} GB"
        else:
            return f"{bytes_value / (1024 * 1024 * 1024 * 1024):.2f} TB"
    
    @staticmethod
    def _format_bytes_to_speed(bytes_per_second):
        """将字节/秒转换为可读的速度字符串"""
        if bytes_per_second < 1024:
            return f"{bytes_per_second} B/s"
        elif bytes_per_second < 1024 * 1024:
            return f"{bytes_per_second / 1024:.2f} KB/s"
        elif bytes_per_second < 1024 * 1024 * 1024:
            return f"{bytes_per_second / (1024 * 1024):.2f} MB/s"
        else:
            return f"{bytes_per_second / (1024 * 1024 * 1024):.2f} GB/s"

